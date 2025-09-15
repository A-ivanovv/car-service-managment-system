from django.core.management.base import BaseCommand
from dashboard.models import Sklad
import openpyxl
from decimal import Decimal


class Command(BaseCommand):
    help = 'Import Sklad data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='НАЛИЧНОСТИ 04.09.2025Г.xlsx',
            help='Excel file path to import from'
        )
        parser.add_argument(
            '--start-row',
            type=int,
            default=8,
            help='Starting row number for data (default: 8)'
        )
        parser.add_argument(
            '--clear-existing',
            action='store_true',
            help='Clear existing data before import'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        start_row = options['start_row']
        clear_existing = options['clear_existing']

        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            self.stdout.write(f"Loading data from: {file_path}")
            self.stdout.write(f"Sheet: {ws.title}")
            self.stdout.write(f"Total rows: {ws.max_row}")
            self.stdout.write(f"Starting from row: {start_row}")

            # Clear existing data if requested
            if clear_existing:
                deleted_count = Sklad.objects.count()
                Sklad.objects.all().delete()
                self.stdout.write(f"Cleared {deleted_count} existing records")

            created_count = 0
            updated_count = 0
            error_count = 0

            # Process each row
            for row_num in range(start_row, ws.max_row + 1):
                try:
                    # Get cell values
                    article_number = ws.cell(row=row_num, column=1).value
                    name = ws.cell(row=row_num, column=2).value
                    unit = ws.cell(row=row_num, column=6).value
                    quantity = ws.cell(row=row_num, column=7).value
                    purchase_price = ws.cell(row=row_num, column=8).value
                    total_value = ws.cell(row=row_num, column=9).value

                    # Skip empty rows
                    if not article_number or not name:
                        continue

                    # Convert to string and clean up
                    article_number = str(article_number).strip()
                    name = str(name).strip()
                    unit = str(unit).strip() if unit else 'бр'
                    
                    # Convert numeric values
                    try:
                        quantity = float(quantity) if quantity is not None else 0.0
                        purchase_price = float(purchase_price) if purchase_price is not None else 0.0
                    except (ValueError, TypeError):
                        self.stdout.write(
                            self.style.WARNING(f"Row {row_num}: Invalid numeric values, skipping")
                        )
                        error_count += 1
                        continue

                    # Check if item already exists
                    existing_item = Sklad.objects.filter(article_number=article_number).first()
                    
                    if existing_item:
                        # Update existing item
                        existing_item.name = name
                        existing_item.unit = unit
                        existing_item.quantity = quantity
                        existing_item.purchase_price = purchase_price
                        existing_item.is_active = True
                        existing_item.save()
                        updated_count += 1
                        self.stdout.write(f"Updated: {article_number} - {name}")
                    else:
                        # Create new item
                        Sklad.objects.create(
                            article_number=article_number,
                            name=name,
                            unit=unit,
                            quantity=quantity,
                            purchase_price=purchase_price,
                            is_active=True
                        )
                        created_count += 1
                        self.stdout.write(f"Created: {article_number} - {name}")

                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(f"Row {row_num}: Error processing - {str(e)}")
                    )
                    error_count += 1
                    continue

            # Summary
            self.stdout.write(
                self.style.SUCCESS(
                    f"\nImport completed!\n"
                    f"Created: {created_count}\n"
                    f"Updated: {updated_count}\n"
                    f"Errors: {error_count}\n"
                    f"Total processed: {created_count + updated_count}"
                )
            )

        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f"File not found: {file_path}")
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f"Error loading Excel file: {str(e)}")
            )
